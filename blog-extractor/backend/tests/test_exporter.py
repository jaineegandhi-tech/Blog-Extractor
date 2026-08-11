import csv
import os

import pytest
from openpyxl import load_workbook

from app.database import Base
from app.models import CrawlJob, BlogResult, FailedURL, JobStatus
from app.services import exporter
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker


@pytest.fixture()
def db_session(tmp_path, monkeypatch):
    engine = create_engine(f"sqlite:///{tmp_path}/test.db")
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    monkeypatch.setattr(exporter.settings, "EXPORT_DIR", str(tmp_path))
    session = Session()
    yield session
    session.close()


def _make_job(db):
    job = CrawlJob(website_url="https://example.com", normalized_domain="example.com",
                    status=JobStatus.COMPLETED)
    db.add(job)
    db.commit()
    db.refresh(job)
    db.add(BlogResult(
        job_id=job.id, website="https://example.com", blog_url="https://example.com/blog/a",
        blog_title="Post A", word_count=500, extraction_status="Success", http_status=200,
    ))
    db.add(BlogResult(
        job_id=job.id, website="https://example.com", blog_url="https://example.com/blog/b",
        blog_title="Post B", word_count=750, extraction_status="Success", http_status=200,
    ))
    db.add(FailedURL(job_id=job.id, url="https://example.com/blog/broken", error="Timeout",
                      http_status=0, retry_count=3))
    db.commit()
    return job


def test_export_xlsx_creates_file_with_rows(db_session):
    job = _make_job(db_session)
    path = exporter.export_xlsx(db_session, job)
    assert os.path.exists(path)
    wb = load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0][4] == "Blog Title"  # header row
    titles = {row[4] for row in rows[1:]}
    assert titles == {"Post A", "Post B"}


def test_export_csv_creates_file_with_rows(db_session):
    job = _make_job(db_session)
    path = exporter.export_csv(db_session, job)
    with open(path, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    assert rows[0][4] == "Blog Title"
    assert len(rows) == 3  # header + 2 results


def test_export_failed_urls_csv(db_session):
    job = _make_job(db_session)
    path = exporter.export_failed_urls_csv(db_session, job)
    with open(path, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    assert rows[0] == ["URL", "Error", "HTTP Status", "Retry Count", "Timestamp"]
    assert rows[1][0] == "https://example.com/blog/broken"
