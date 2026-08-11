"""
Excel / CSV export (§7, §17, §20). Streams rows using openpyxl's write-only
mode so multi-thousand-row exports don't blow up memory.
"""
import csv
import os
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.orm import Session

from app.config import settings
from app.models import BlogResult, FailedURL, CrawlJob
from app.services.url_utils import get_domain

COLUMNS = [
    ("id_", "ID", 6),
    ("website", "Website", 20),
    ("blog_url", "Blog URL", 40),
    ("canonical_url", "Canonical URL", 40),
    ("blog_title", "Blog Title", 30),
    ("meta_title", "Meta Title", 30),
    ("meta_description", "Meta Description", 40),
    ("author", "Author", 18),
    ("author_url", "Author URL", 25),
    ("publication_date", "Publication Date", 16),
    ("last_updated_date", "Last Updated Date", 16),
    ("category", "Category", 16),
    ("subcategory", "Subcategory", 16),
    ("tags", "Tags", 25),
    ("featured_image_url", "Featured Image URL", 30),
    ("word_count", "Word Count", 10),
    ("h1", "H1", 30),
    ("h2_headings", "H2 Headings", 40),
    ("h3_headings", "H3 Headings", 40),
    ("main_content", "Main Content", 60),
    ("introduction", "Introduction", 40),
    ("conclusion", "Conclusion", 40),
    ("faq_content", "FAQ Content", 40),
    ("extraction_status", "Extraction Status", 14),
    ("http_status", "HTTP Status", 10),
    ("extraction_date", "Extraction Date", 18),
    ("error_message", "Error Message", 30),
]


def _row_dict(idx: int, r: BlogResult) -> dict:
    return {
        "id_": idx,
        "website": r.website,
        "blog_url": r.blog_url,
        "canonical_url": r.canonical_url,
        "blog_title": r.blog_title,
        "meta_title": r.meta_title,
        "meta_description": r.meta_description,
        "author": r.author,
        "author_url": r.author_url,
        "publication_date": r.publication_date,
        "last_updated_date": r.last_updated_date,
        "category": r.category,
        "subcategory": r.subcategory,
        "tags": r.tags,
        "featured_image_url": r.featured_image_url,
        "word_count": r.word_count,
        "h1": r.h1,
        "h2_headings": r.h2_headings,
        "h3_headings": r.h3_headings,
        "main_content": r.main_content,
        "introduction": r.introduction,
        "conclusion": r.conclusion,
        "faq_content": r.faq_content,
        "extraction_status": r.extraction_status,
        "http_status": r.http_status,
        "extraction_date": r.extraction_date.isoformat() if r.extraction_date else None,
        "error_message": r.error_message,
    }


def _iter_results(db: Session, job_id: str, batch_size: int = 500):
    offset = 0
    while True:
        batch = (
            db.query(BlogResult)
            .filter(BlogResult.job_id == job_id)
            .order_by(BlogResult.extraction_date)
            .offset(offset)
            .limit(batch_size)
            .all()
        )
        if not batch:
            break
        for r in batch:
            yield r
        offset += batch_size


def export_xlsx(db: Session, job: CrawlJob) -> str:
    os.makedirs(settings.EXPORT_DIR, exist_ok=True)
    domain = get_domain(job.website_url)
    date_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    filename = f"{domain}_blog_extraction_{date_str}.xlsx"
    path = os.path.join(settings.EXPORT_DIR, filename)

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Blog Extraction")

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    wrap = Alignment(wrap_text=True, vertical="top")

    header_cells = []
    for key, label, width in COLUMNS:
        from openpyxl.cell import WriteOnlyCell
        c = WriteOnlyCell(ws, value=label)
        c.font = header_font
        c.fill = header_fill
        header_cells.append(c)
    ws.append(header_cells)

    for idx, r in enumerate(_iter_results(db, job.id), start=1):
        row = _row_dict(idx, r)
        cells = []
        for key, _, _ in COLUMNS:
            from openpyxl.cell import WriteOnlyCell
            c = WriteOnlyCell(ws, value=row[key])
            c.alignment = wrap
            cells.append(c)
        ws.append(cells)

    # Column widths + freeze header row (write-only mode still supports these)
    for i, (_, _, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"

    wb.save(path)
    return path


def export_csv(db: Session, job: CrawlJob) -> str:
    os.makedirs(settings.EXPORT_DIR, exist_ok=True)
    domain = get_domain(job.website_url)
    date_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    filename = f"{domain}_blog_extraction_{date_str}.csv"
    path = os.path.join(settings.EXPORT_DIR, filename)

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([label for _, label, _ in COLUMNS])
        for idx, r in enumerate(_iter_results(db, job.id), start=1):
            row = _row_dict(idx, r)
            writer.writerow([row[key] for key, _, _ in COLUMNS])

    return path


def export_failed_urls_csv(db: Session, job: CrawlJob) -> str:
    os.makedirs(settings.EXPORT_DIR, exist_ok=True)
    path = os.path.join(settings.EXPORT_DIR, f"{job.id}_failed_urls.csv")

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["URL", "Error", "HTTP Status", "Retry Count", "Timestamp"])
        q = db.query(FailedURL).filter(FailedURL.job_id == job.id).order_by(FailedURL.timestamp)
        for fu in q:
            writer.writerow([fu.url, fu.error, fu.http_status, fu.retry_count,
                              fu.timestamp.isoformat() if fu.timestamp else None])
    return path
